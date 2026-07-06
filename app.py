import streamlit as st
import pandas as pd
import requests
import io
import zipfile
import re
import os
import shutil
import openpyxl

st.set_page_config(
    page_title="Gerenciador de Catálogos e PDFs",
    page_icon="📚",
    layout="wide"
)

st.title("📚 Gerenciador de Download de Conteúdos (PDFs)")
st.markdown("---")

# =========================
# Funções auxiliares
# =========================

def limpar_nome_arquivo(nome, limite=140):
    if pd.isna(nome) or str(nome).strip() == "":
        return "sem_nome"

    nome_limpo = re.sub(r'[\\/*?:"<>|]', "", str(nome))
    nome_limpo = re.sub(r"\s+", " ", nome_limpo).strip()

    if len(nome_limpo) > limite:
        nome_limpo = nome_limpo[:limite].strip()

    return nome_limpo if nome_limpo else "sem_nome"


def criar_slug_curso(nome_curso):
    if pd.isna(nome_curso) or str(nome_curso).strip() == "":
        return "curso_sem_nome"

    slug = str(nome_curso).lower().strip()
    slug = re.sub(r"\s+", "_", slug)
    slug = re.sub(r'[\\/*?:"<>|]', "", slug)

    return slug if slug else "curso_sem_nome"


def garantir_nome_unico(caminho_base, nomes_usados):
    """
    Evita arquivos duplicados dentro do ZIP.
    """
    if caminho_base not in nomes_usados:
        nomes_usados[caminho_base] = 1
        return caminho_base

    nomes_usados[caminho_base] += 1

    pasta, arquivo = os.path.split(caminho_base)
    nome, extensao = os.path.splitext(arquivo)

    novo_nome = f"{nome}_{nomes_usados[caminho_base]}{extensao}"
    return os.path.join(pasta, novo_nome).replace("\\", "/")


def carregar_dados_otimizados(nome_arquivo, arquivo_bytes):
    """
    Lê o arquivo carregado de forma mais estável.
    Evita usar o objeto UploadedFile diretamente com cache.
    """

    colunas_obrigatorias = [
        "ANO",
        "PERÍODO",
        "CURSO",
        "UC",
        "CATÁLOGO",
        "ORDEM AULA",
        "ORDEM ATIVIDADE",
        "ATIVIDADE",
        "PDF"
    ]

    extensao = nome_arquivo.lower().split(".")[-1]

    if extensao == "csv":
        buffer = io.BytesIO(arquivo_bytes)

        df = pd.read_csv(
            buffer,
            dtype=str,
            encoding="utf-8",
            sep=None,
            engine="python"
        )

        df.columns = [str(col).strip() for col in df.columns]

        faltam = [col for col in colunas_obrigatorias if col not in df.columns]
        if faltam:
            raise ValueError(
                f"Faltam as seguintes colunas obrigatórias: {', '.join(faltam)}"
            )

        df = df[colunas_obrigatorias].copy()

    elif extensao in ["xlsx", "xlsm"]:
        buffer = io.BytesIO(arquivo_bytes)

        wb = openpyxl.load_workbook(
            buffer,
            read_only=True,
            data_only=True
        )

        ws = wb.active

        header = []
        for cell in ws[1]:
            valor = "" if cell.value is None else str(cell.value).strip()
            header.append(valor)

        col_indices = {}

        for col in colunas_obrigatorias:
            if col in header:
                col_indices[col] = header.index(col)

        faltam = [col for col in colunas_obrigatorias if col not in col_indices]
        if faltam:
            raise ValueError(
                f"Faltam as seguintes colunas obrigatórias: {', '.join(faltam)}"
            )

        data = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue

            linha_filtrada = {}

            for col, idx in col_indices.items():
                valor = row[idx] if idx < len(row) else ""
                linha_filtrada[col] = "" if valor is None else str(valor)

            data.append(linha_filtrada)

        df = pd.DataFrame(data)

        wb.close()

    else:
        raise ValueError(
            "Formato não suportado. Use CSV, XLSX ou XLSM."
        )

    df.columns = [str(col).strip() for col in df.columns]

    for col in colunas_obrigatorias:
        df[col] = df[col].fillna("").astype(str).str.strip()

    return df


def baixar_pdf_para_zip(session, zip_file, url, caminho_no_zip):
    """
    Baixa o PDF em streaming e grava diretamente no ZIP,
    sem carregar o arquivo inteiro em memória.
    """

    headers = {
        "User-Agent": "Mozilla/5.0"
    }

    response = session.get(
        url,
        timeout=(10, 60),
        stream=True,
        allow_redirects=True,
        headers=headers
    )

    if response.status_code != 200:
        raise Exception(f"Erro HTTP {response.status_code}")

    content_type = response.headers.get("Content-Type", "").lower()

    if "text/html" in content_type:
        raise Exception("O link retornou HTML em vez de PDF")

    with zip_file.open(caminho_no_zip, "w") as arquivo_zip:
        for chunk in response.iter_content(chunk_size=1024 * 128):
            if chunk:
                arquivo_zip.write(chunk)


# =========================
# Pastas temporárias
# =========================

PASTA_TMP = "/tmp/download_workspace"
ARQUIVO_ZIP_FINAL = "/tmp/resultado_curso.zip"

# =========================
# Sidebar
# =========================

st.sidebar.header("📁 Upload da Base de Dados")

arquivo_carregado = st.sidebar.file_uploader(
    "Carregue o arquivo unificado (.xlsx, .xlsm ou .csv)",
    type=["xlsx", "xlsm", "csv"]
)

if arquivo_carregado is None:
    st.session_state["zip_disponivel_no_disco"] = False
    st.info("💡 Por favor, carregue o arquivo Excel ou CSV na barra lateral.")
    st.stop()

try:
    arquivo_bytes = arquivo_carregado.getvalue()

    with st.spinner("Processando arquivo em modo estável..."):
        df = carregar_dados_otimizados(
            arquivo_carregado.name,
            arquivo_bytes
        )

    st.success("✅ Base de dados carregada com sucesso!")

    # =========================
    # Filtros
    # =========================

    st.subheader("🔍 Filtros de Seleção")

    col1, col2, col3 = st.columns(3)

    with col1:
        anos_disponiveis = sorted(df["ANO"].dropna().unique())

        if len(anos_disponiveis) == 0:
            st.warning("Nenhum ANO encontrado na base.")
            st.stop()

        ano_selecionado = st.selectbox(
            "1. Selecione o ANO:",
            anos_disponiveis
        )

    df_ano = df[df["ANO"] == ano_selecionado]

    with col2:
        periodos_disponiveis = sorted(df_ano["PERÍODO"].dropna().unique())

        if len(periodos_disponiveis) == 0:
            st.warning("Nenhum PERÍODO encontrado para o ano selecionado.")
            st.stop()

        periodo_selecionado = st.selectbox(
            "2. Selecione o PERÍODO:",
            periodos_disponiveis
        )

    df_periodo = df_ano[df_ano["PERÍODO"] == periodo_selecionado]

    with col3:
        cursos_disponiveis = sorted(df_periodo["CURSO"].dropna().unique())

        if len(cursos_disponiveis) == 0:
            st.warning("Nenhum CURSO encontrado para o período selecionado.")
            st.stop()

        curso_selecionado = st.selectbox(
            "3. Selecione o CURSO:",
            cursos_disponiveis
        )

    df_filtrado = df_periodo[df_periodo["CURSO"] == curso_selecionado].copy()

    st.markdown("---")

    # =========================
    # Resumo
    # =========================

    st.subheader("📊 Resumo dos Dados Disponíveis")

    qtd_uc = df_filtrado["UC"].nunique()
    qtd_catalogo = df_filtrado["CATÁLOGO"].nunique()
    total_linhas = len(df_filtrado)

    m1, m2, m3 = st.columns(3)
    m1.metric("Quantidade de UCs Únicas", qtd_uc)
    m2.metric("Quantidade de Catálogos Únicos", qtd_catalogo)
    m3.metric("Total de Arquivos/Aulas", total_linhas)

    with st.expander("👀 Visualizar linhas que serão processadas"):
        st.dataframe(df_filtrado, use_container_width=True)

    st.markdown("---")

    # =========================
    # Processamento
    # =========================

    st.subheader("⚙️ Processamento e Download dos PDFs")

    with st.form(key="form_processamento"):
        st.write("Clique abaixo para processar os PDFs e gerar o arquivo ZIP.")
        botao_disparar = st.form_submit_button(
            "🚀 Iniciar Processamento",
            type="primary"
        )

    if botao_disparar:
        if os.path.exists(PASTA_TMP):
            shutil.rmtree(PASTA_TMP, ignore_errors=True)

        if os.path.exists(ARQUIVO_ZIP_FINAL):
            os.remove(ARQUIVO_ZIP_FINAL)

        df_com_pdf = df_filtrado.copy()
        df_com_pdf["PDF_STR"] = df_com_pdf["PDF"].astype(str).str.strip()

        df_com_pdf = df_com_pdf[
            (df_com_pdf["PDF_STR"] != "") &
            (df_com_pdf["PDF_STR"].str.startswith("http"))
        ].copy()

        if len(df_com_pdf) == 0:
            st.warning("⚠️ Nenhum PDF válido encontrado.")
            st.session_state["zip_disponivel_no_disco"] = False
            st.stop()

        progresso_bar = st.progress(0)
        status_text = st.empty()

        nome_pasta_curso = criar_slug_curso(curso_selecionado)

        sucessos = 0
        erros = 0
        total = len(df_com_pdf)

        relatorio_erros = []
        nomes_usados = {}

        session = requests.Session()

        try:
            with zipfile.ZipFile(
                ARQUIVO_ZIP_FINAL,
                "w",
                compression=zipfile.ZIP_DEFLATED,
                allowZip64=True
            ) as zip_file:

                lista_dados = df_com_pdf.to_dict(orient="records")

                for idx, row in enumerate(lista_dados):
                    porcentagem = (idx + 1) / total
                    progresso_bar.progress(porcentagem)

                    atividade_nome = str(row.get("ATIVIDADE", "Sem Nome"))
                    uc_nome = str(row.get("UC", "Sem_UC"))
                    link_pdf = row.get("PDF_STR", "")

                    status_text.text(
                        f"Baixando {idx + 1}/{total}: {atividade_nome}"
                    )

                    try:
                        ordem_aula = limpar_nome_arquivo(
                            row.get("ORDEM AULA", "0"),
                            limite=30
                        )

                        ordem_atv = limpar_nome_arquivo(
                            row.get("ORDEM ATIVIDADE", "0"),
                            limite=30
                        )

                        atv = limpar_nome_arquivo(
                            atividade_nome,
                            limite=80
                        )

                        uc = limpar_nome_arquivo(
                            uc_nome,
                            limite=80
                        )

                        nome_pdf = f"{ordem_aula}.{ordem_atv} - {uc} - {atv}.pdf"

                        caminho_no_zip = (
                            f"{nome_pasta_curso}/"
                            f"04_conteudos_especificos/"
                            f"conteudo/"
                            f"{nome_pdf}"
                        )

                        caminho_no_zip = garantir_nome_unico(
                            caminho_no_zip,
                            nomes_usados
                        )

                        baixar_pdf_para_zip(
                            session=session,
                            zip_file=zip_file,
                            url=link_pdf,
                            caminho_no_zip=caminho_no_zip
                        )

                        sucessos += 1

                    except Exception as e:
                        erros += 1

                        relatorio_erros.append({
                            "UC": uc_nome,
                            "Atividade": atividade_nome,
                            "Motivo": str(e),
                            "Link": link_pdf
                        })

                        continue

            status_text.text("✨ Compactação concluída!")

        except Exception as e:
            st.session_state["zip_disponivel_no_disco"] = False
            st.error(f"Erro ao criar o arquivo ZIP: {e}")
            st.stop()

        if sucessos > 0:
            st.success(f"📊 {sucessos} PDFs foram estruturados com sucesso!")
            st.session_state["zip_disponivel_no_disco"] = True
            st.session_state["nome_do_curso_zip"] = nome_pasta_curso
        else:
            st.session_state["zip_disponivel_no_disco"] = False
            st.error("Nenhum PDF foi baixado com sucesso.")

        if erros > 0:
            st.warning(f"⚠️ {erros} links apresentaram falhas.")

            with st.expander("🚨 Ver detalhes das falhas"):
                df_erros = pd.DataFrame(relatorio_erros)
                st.dataframe(df_erros, use_container_width=True)

    # =========================
    # Download do ZIP
    # =========================

    if (
        st.session_state.get("zip_disponivel_no_disco")
        and os.path.exists(ARQUIVO_ZIP_FINAL)
    ):
        st.markdown("### 📥 Seu arquivo está pronto:")

        nome_download = f"{st.session_state.get('nome_do_curso_zip')}.zip"

        tamanho_mb = os.path.getsize(ARQUIVO_ZIP_FINAL) / (1024 * 1024)
        st.caption(f"Tamanho aproximado do arquivo: {tamanho_mb:.2f} MB")

        if tamanho_mb > 200:
            st.warning(
                "⚠️ O ZIP ficou muito grande. O botão de download pode demorar "
                "ou falhar dependendo do limite de memória do ambiente Streamlit."
            )

        with open(ARQUIVO_ZIP_FINAL, "rb") as arquivo_zip:
            dados_zip = arquivo_zip.read()

        st.download_button(
            label=f"Baixar {nome_download}",
            data=dados_zip,
            file_name=nome_download,
            mime="application/zip",
            use_container_width=True
        )

except Exception as e:
    st.session_state["zip_disponivel_no_disco"] = False
    st.error(f"Erro crítico: {e}")
